#!/usr/bin/env python3
import base64, csv, hashlib, hmac, io, json, os, secrets, socketserver, time, urllib.parse, urllib.request
from collections import Counter, deque
from datetime import datetime, timezone
from html import escape
import http.server

TOK = os.environ.get("GH_WORKFLOW_TOKEN", "")
REPO = os.environ.get("GH_REPO", "Custimoo-Ops/custimoo-defect-report")
PORT = int(os.environ.get("PORT", 8080))
DQC_URL = os.environ.get("DQC_EVENTS_URL", "https://dqc-dashboard-custimoo.fly.dev/api/v1/dqc")
DQC_API_KEY = os.environ.get("DQC_API_KEY", "")
# Legacy Basic auth fallback kept for older deployments only.
DQC_USER = os.environ.get("DQC_DASH_USER", "")
DQC_PASS = os.environ.get("DQC_DASH_PASSWORD", "")
DQC_SKILL_VERSION = os.environ.get("DQC_SKILL_VERSION", "0.5.5")

SSO_ENABLED = os.environ.get("CUSTIMOO_SSO_ENABLED", "1").lower() not in ("0", "false", "no", "off")
SSO_TENANT_ID = os.environ.get("CUSTIMOO_SSO_TENANT_ID") or os.environ.get("CUSTIMOO_GRAPH_TENANT_ID", "")
SSO_CLIENT_ID = os.environ.get("CUSTIMOO_SSO_CLIENT_ID") or os.environ.get("CUSTIMOO_GRAPH_CLIENT_ID", "")
SSO_CLIENT_SECRET = os.environ.get("CUSTIMOO_SSO_CLIENT_SECRET") or os.environ.get("CUSTIMOO_GRAPH_CLIENT_SECRET", "")
SSO_SESSION_SECRET = os.environ.get("CUSTIMOO_SSO_SESSION_SECRET", "")
SSO_ALLOWED_DOMAIN = os.environ.get("CUSTIMOO_SSO_ALLOWED_DOMAIN", "custimoo.com").lower().lstrip("@")
SSO_COOKIE = "custimoo_report_session"
SSO_STATE_COOKIE = "custimoo_report_oauth_state"
SSO_NONCE_COOKIE = "custimoo_report_oauth_nonce"
SSO_SESSION_TTL = int(os.environ.get("CUSTIMOO_SSO_SESSION_TTL", str(12 * 3600)))
VISITS = deque(maxlen=int(os.environ.get("VISIT_LOG_MAX", "5000")))
def b64url_decode(segment):
    segment += "=" * (-len(segment) % 4)
    return base64.urlsafe_b64decode(segment.encode())

def b64url_json(segment):
    return json.loads(b64url_decode(segment).decode())

def cookie_header(name, value, max_age=None):
    parts = [f"{name}={value}", "Path=/", "HttpOnly", "Secure", "SameSite=Lax"]
    if max_age is not None:
        parts.append(f"Max-Age={int(max_age)}")
    return "; ".join(parts)

def clear_cookie_header(name):
    return cookie_header(name, "", 0)

def sso_configured():
    return all([SSO_TENANT_ID, SSO_CLIENT_ID, SSO_CLIENT_SECRET, SSO_SESSION_SECRET])

REASON_KEYS = ("rejection_reason", "reject_reason", "reason", "failure_reason", "qc_reason", "notes", "message")

def event_reason(e):
    for k in REASON_KEYS:
        v = e.get(k)
        if v is not None and str(v).strip():
            return str(v).strip()
    friction = e.get("friction") if isinstance(e.get("friction"), dict) else {}
    if friction and friction.get("has_friction") is False:
        return ""
    for k in ("limitation", "note", "status"):
        v = friction.get(k)
        if v is not None and str(v).strip() and str(v).strip().lower() not in ("none", "ignored", "non-friction"):
            return str(v).strip()
    return ""

USER_KEYS = ("windows_login", "windows_user", "windows_username", "login_name", "username", "user")

def event_user(e):
    reviewer = e.get("reviewer") if isinstance(e.get("reviewer"), dict) else {}
    for v in (reviewer.get("name"), e.get("display_user")):
        if v is not None and str(v).strip():
            return str(v).strip()
    for k in USER_KEYS:
        v = e.get(k)
        if v is not None and str(v).strip():
            return str(v).strip()
    return "(unknown)"

def event_version(e):
    return str(e.get("dqc_skill_version") or e.get("version") or DQC_SKILL_VERSION or "").strip()

def normalize_dqc_event(e):
    out = dict(e)
    out["display_user"] = event_user(out)
    out["user"] = out.get("user") or out["display_user"]
    out["ts"] = str(out.get("ts") or out.get("timestamp_utc") or out.get("created_at") or "")
    out["date"] = str(out.get("date") or out["ts"][:10])
    out["verdict"] = str(out.get("verdict") or out.get("status") or "UNKNOWN").upper()
    out["order"] = str(out.get("order") or out.get("order_no") or "")
    out["rejection_reason"] = event_reason(out)
    out["dqc_skill_version"] = event_version(out)
    return out

DQC_PAGE = """<!doctype html>
<html><head><meta charset="utf-8"><meta name="viewport" content="width=device-width, initial-scale=1">
<title>Custimoo DQC Usage</title>
<style>
body{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;background:#f5f7fb;color:#172033;margin:0;padding:24px}.wrap{max-width:1200px;margin:auto}.top{display:flex;gap:12px;align-items:center;justify-content:space-between;flex-wrap:wrap}.card{background:#fff;border:1px solid #e6eaf2;border-radius:14px;padding:18px;margin:14px 0;box-shadow:0 8px 24px rgba(16,24,40,.06)}h1{margin:0 0 4px}.muted{color:#667085;font-size:13px}.filters{display:flex;gap:10px;flex-wrap:wrap;align-items:end}.filters label{font-size:12px;font-weight:700;color:#475467;display:block;margin-bottom:4px}input,button{padding:9px 10px;border:1px solid #d0d5dd;border-radius:8px;background:white}button{background:#0f3460;color:white;cursor:pointer;border-color:#0f3460}button.secondary{background:white;color:#0f3460}.grid{display:grid;grid-template-columns:repeat(4,minmax(140px,1fr));gap:12px}.kpi .label{font-size:12px;color:#667085;text-transform:uppercase;font-weight:700}.kpi .value{font-size:30px;font-weight:800;margin-top:4px}.passed{color:#07864b}.rejected{color:#c62828}table{width:100%;border-collapse:collapse;font-size:13px}th,td{padding:9px;border-bottom:1px solid #eef2f7;text-align:left}th{background:#f8fafc;font-weight:800}.right{text-align:right}.pill{display:inline-block;padding:3px 8px;border-radius:999px;font-weight:700;font-size:12px}.pill.PASSED{background:#e8f5ee;color:#087443}.pill.REJECTED{background:#fdeaea;color:#b42318}.error{color:#b42318;font-weight:700}@media(max-width:760px){.grid{grid-template-columns:1fr 1fr}}
</style></head><body><div class="wrap">
<div class="top"><div><h1>Custimoo Digital QC Usage</h1><div class="muted" id="generated">Loading…</div></div><div><a href="/" class="muted">← Failure report</a></div></div>
<div class="card filters"><div><label>From</label><input id="from" type="date"></div><div><label>To</label><input id="to" type="date"></div><div><button onclick="loadData()">Refresh</button></div><div><button class="secondary" onclick="download('/api/dqc.csv')">Download CSV</button></div><div><button class="secondary" onclick="download('/api/dqc.xlsx')">Download Excel</button></div></div>
<div id="msg" class="muted"></div>
<div class="grid"><div class="card kpi"><div class="label">Total audits</div><div class="value" id="total">–</div></div><div class="card kpi"><div class="label">PASSED</div><div class="value passed" id="passed">–</div></div><div class="card kpi"><div class="label">REJECTED</div><div class="value rejected" id="rejected">–</div></div><div class="card kpi"><div class="label">Users</div><div class="value" id="users">–</div></div></div>
<div class="card"><h3>Per-user count</h3><table><thead><tr><th>User</th><th class="right">Audits</th></tr></thead><tbody id="userBody"></tbody></table></div>
<div class="card"><h3>All runs</h3><table><thead><tr><th>Date</th><th>User</th><th>Order</th><th>Verdict</th><th>Rejection reason</th><th>DQC Skill Version</th><th>Timestamp UTC</th></tr></thead><tbody id="runBody"></tbody></table></div>
</div><script>
function qs(){const p=new URLSearchParams(); const f=document.getElementById('from').value,t=document.getElementById('to').value; if(f)p.set('from',f); if(t)p.set('to',t); return p.toString()?('?'+p.toString()):''}
function download(path){location.href=path+qs()}
function esc(v){return String(v==null?'':v).replace(/[&<>"']/g,c=>({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}[c]))}
async function loadData(){document.getElementById('msg').textContent='Loading…'; try{const r=await fetch('/api/dqc/events'+qs()); const d=await r.json(); if(!r.ok) throw new Error(d.error||r.statusText); render(d)}catch(e){document.getElementById('msg').innerHTML='<span class="error">'+esc(e.message)+'</span>'}}
function render(d){const ev=d.events||[]; document.getElementById('generated').textContent='API generated: '+(d.generated_at||'n/a')+' · '+ev.length+' rows'; document.getElementById('msg').textContent=d.stale_error?('Warning: '+d.stale_error):''; const vc={PASSED:0,REJECTED:0}; const uc={}; const reason=e=>(e.rejection_reason||e.reject_reason||e.reason||e.failure_reason||e.qc_reason||e.notes||e.message||'—'); const user=e=>{const r=e.reviewer&&typeof e.reviewer==='object'?e.reviewer:{}; return e.display_user||r.name||e.windows_login||e.windows_user||e.windows_username||e.login_name||e.username||e.user||'(unknown)'}; const ver=e=>(e.dqc_skill_version||e.version||''); const verdict=e=>{let v=String(e.verdict||e.status||'UNKNOWN').toUpperCase(); if(v==='PASS')v='PASSED'; if(v==='FAIL'||v==='FAILED')v='REJECTED'; return v}; ev.forEach(e=>{const v=verdict(e); vc[v]=(vc[v]||0)+1; uc[user(e)]=(uc[user(e)]||0)+1}); document.getElementById('total').textContent=ev.length; document.getElementById('passed').textContent=vc.PASSED||0; document.getElementById('rejected').textContent=vc.REJECTED||0; document.getElementById('users').textContent=Object.keys(uc).length; document.getElementById('userBody').innerHTML=Object.entries(uc).sort((a,b)=>b[1]-a[1]).map(([u,c])=>`<tr><td>${esc(u)}</td><td class="right">${c}</td></tr>`).join('')||'<tr><td colspan=2>No users</td></tr>'; document.getElementById('runBody').innerHTML=ev.map(e=>`<tr><td>${esc((e.ts||'').slice(0,10))}</td><td>${esc(user(e))}</td><td>${esc(e.order||e.order_no||'')}</td><td><span class="pill ${esc(verdict(e))}">${esc(verdict(e))}</span></td><td>${esc(reason(e))}</td><td>${esc(ver(e))}</td><td>${esc(e.ts||'')}</td></tr>`).join('')||'<tr><td colspan=7>No audits logged</td></tr>'}
loadData();
</script></body></html>"""

class H(http.server.SimpleHTTPRequestHandler):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, directory="/app", **kwargs)

    def _cookies(self):
        out = {}
        for part in (self.headers.get("Cookie") or "").split(";"):
            if "=" in part:
                k, v = part.strip().split("=", 1)
                out[k] = urllib.parse.unquote(v)
        return out

    def _external_base_url(self):
        proto = self.headers.get("X-Forwarded-Proto") or "https"
        host = self.headers.get("X-Forwarded-Host") or self.headers.get("Host") or "custimoo-defect-report-lars.fly.dev"
        return f"{proto}://{host}"

    def _redirect_uri(self):
        return self._external_base_url() + "/auth/callback"

    def _sign(self, payload):
        raw = json.dumps(payload, separators=(",", ":"), sort_keys=True).encode()
        data = base64.urlsafe_b64encode(raw).decode().rstrip("=")
        sig = hmac.new(SSO_SESSION_SECRET.encode(), data.encode(), hashlib.sha256).digest()
        return data + "." + base64.urlsafe_b64encode(sig).decode().rstrip("=")

    def _unsign(self, token):
        try:
            data, sig = token.split(".", 1)
            expected = hmac.new(SSO_SESSION_SECRET.encode(), data.encode(), hashlib.sha256).digest()
            actual = b64url_decode(sig)
            if not hmac.compare_digest(expected, actual):
                return None
            payload = json.loads(b64url_decode(data).decode())
            if int(payload.get("exp", 0)) < int(time.time()):
                return None
            return payload
        except Exception:
            return None

    def _current_user(self):
        if not SSO_ENABLED:
            return {"email": "sso-disabled"}
        if not sso_configured():
            return None
        return self._unsign(self._cookies().get(SSO_COOKIE, ""))

    def _is_public_path(self, path):
        return path.startswith("/auth/") or path in ("/favicon.ico",)

    def _require_auth(self, path):
        if not SSO_ENABLED or self._is_public_path(path):
            return True
        if not sso_configured():
            self._html("""<!doctype html><title>SSO setup required</title><body style='font-family:sans-serif;padding:32px'><h1>Custimoo SSO is not configured</h1><p>Missing one or more server secrets: tenant id, client id, client secret, or session secret.</p></body>""", code=503)
            return False
        if self._current_user():
            return True
        next_path = self.path if self.path.startswith("/") else "/"
        return self._redirect("/auth/login?next=" + urllib.parse.quote(next_path, safe=""))

    def _auth_login(self):
        if not sso_configured():
            return self._json(503, {"error": "Custimoo SSO is not configured"})
        query = urllib.parse.parse_qs(urllib.parse.urlparse(self.path).query)
        next_path = query.get("next", ["/"])[0]
        if not next_path.startswith("/") or next_path.startswith("//"):
            next_path = "/"
        state = secrets.token_urlsafe(24)
        nonce = secrets.token_urlsafe(24)
        params = {
            "client_id": SSO_CLIENT_ID,
            "response_type": "code",
            "redirect_uri": self._redirect_uri(),
            "response_mode": "query",
            "scope": "openid email profile",
            "state": state,
            "nonce": nonce,
            "prompt": "select_account",
        }
        self.send_response(302)
        self.send_header("Set-Cookie", cookie_header(SSO_STATE_COOKIE, self._sign({"state": state, "next": next_path, "exp": int(time.time()) + 600}), 600))
        self.send_header("Set-Cookie", cookie_header(SSO_NONCE_COOKIE, nonce, 600))
        self.send_header("Location", f"https://login.microsoftonline.com/{SSO_TENANT_ID}/oauth2/v2.0/authorize?" + urllib.parse.urlencode(params))
        self.end_headers()

    def _auth_callback(self):
        query = urllib.parse.parse_qs(urllib.parse.urlparse(self.path).query)
        if query.get("error"):
            return self._html("<h1>Custimoo SSO login failed</h1><pre>%s</pre>" % escape(str(query)), code=401)
        code = query.get("code", [""])[0]
        state = query.get("state", [""])[0]
        state_payload = self._unsign(self._cookies().get(SSO_STATE_COOKIE, ""))
        nonce = self._cookies().get(SSO_NONCE_COOKIE, "")
        if not code or not state_payload or state_payload.get("state") != state or not nonce:
            return self._html("<h1>Invalid or expired SSO state</h1><p>Please try logging in again.</p>", code=401)
        body = urllib.parse.urlencode({
            "client_id": SSO_CLIENT_ID,
            "client_secret": SSO_CLIENT_SECRET,
            "grant_type": "authorization_code",
            "code": code,
            "redirect_uri": self._redirect_uri(),
            "scope": "openid email profile",
        }).encode()
        req = urllib.request.Request(
            f"https://login.microsoftonline.com/{SSO_TENANT_ID}/oauth2/v2.0/token",
            data=body,
            method="POST",
            headers={"Content-Type": "application/x-www-form-urlencoded", "Accept": "application/json"},
        )
        try:
            with urllib.request.urlopen(req, timeout=20) as r:
                token = json.loads(r.read().decode())
        except Exception as e:
            return self._html("<h1>Custimoo SSO token exchange failed</h1><pre>%s</pre>" % escape(str(e)[:500]), code=401)
        id_token = token.get("id_token", "")
        try:
            claims = b64url_json(id_token.split(".")[1])
        except Exception:
            return self._html("<h1>Invalid SSO token</h1>", code=401)
        now = int(time.time())
        email = (claims.get("preferred_username") or claims.get("email") or claims.get("upn") or "").lower()
        if claims.get("aud") != SSO_CLIENT_ID or int(claims.get("exp", 0)) < now or claims.get("nonce") != nonce:
            return self._html("<h1>SSO token validation failed</h1>", code=401)
        if SSO_ALLOWED_DOMAIN and not email.endswith("@" + SSO_ALLOWED_DOMAIN):
            return self._html("<h1>Access denied</h1><p>Use your Custimoo account.</p>", code=403)
        session = self._sign({"email": email, "name": claims.get("name") or email, "exp": now + SSO_SESSION_TTL})
        self.send_response(302)
        self.send_header("Set-Cookie", cookie_header(SSO_COOKIE, session, SSO_SESSION_TTL))
        self.send_header("Set-Cookie", clear_cookie_header(SSO_STATE_COOKIE))
        self.send_header("Set-Cookie", clear_cookie_header(SSO_NONCE_COOKIE))
        self.send_header("Location", state_payload.get("next") or "/")
        self.end_headers()

    def _auth_logout(self):
        self.send_response(302)
        self.send_header("Set-Cookie", clear_cookie_header(SSO_COOKIE))
        self.send_header("Location", "/auth/login")
        self.end_headers()

    def _redirect(self, location):
        self.send_response(302)
        self.send_header("Location", location)
        self.end_headers()

    def _client_ip(self):
        return (self.headers.get("Fly-Client-IP") or self.headers.get("X-Forwarded-For") or self.client_address[0] or "").split(",")[0].strip()

    def _track_visit(self, path):
        user = self._current_user() or {}
        event = {
            "ts": datetime.now(timezone.utc).isoformat(),
            "email": user.get("email") or "unknown",
            "name": user.get("name") or user.get("email") or "unknown",
            "path": path,
            "ip": self._client_ip(),
            "user_agent": (self.headers.get("User-Agent") or "")[:180],
        }
        VISITS.appendleft(event)
        print(json.dumps({"event": "report_visit", **event}, separators=(",", ":")), flush=True)

    def _visits(self):
        events = list(VISITS)
        by_user = Counter(e.get("email") or "unknown" for e in events)
        by_path = Counter(e.get("path") or "" for e in events)
        return self._json(200, {
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "retention": "in-memory last %d visits plus Fly app logs" % VISITS.maxlen,
            "total_in_memory": len(events),
            "by_user": dict(by_user.most_common()),
            "by_path": dict(by_path.most_common()),
            "recent": events[:200],
        })

    def do_GET(self):
        path = urllib.parse.urlparse(self.path).path
        if path == "/auth/login": return self._auth_login()
        if path == "/auth/callback": return self._auth_callback()
        if path == "/auth/logout": return self._auth_logout()
        if not self._require_auth(path): return
        if path == "/api/refresh": return self._refresh()
        if path == "/api/status": return self._status()
        if path == "/api/visits": return self._visits()
        if path == "/api/dqc/events": return self._dqc_events()
        if path == "/api/dqc.csv": return self._dqc_csv()
        if path == "/api/dqc.xlsx": return self._dqc_xlsx()
        if path == "/dqc":
            self._track_visit(path)
            return self._html(DQC_PAGE)
        if path == "/":
            self._track_visit(path)
            self.path = "/index.html"
            return super().do_GET()
        return super().do_GET()

    def _github_call(self, url, method="GET", body=None):
        headers = {"Authorization": "Bearer " + TOK, "Accept": "application/vnd.github+json"}
        if body: headers["Content-Type"] = "application/json"
        req = urllib.request.Request(url, data=body, method=method, headers=headers)
        with urllib.request.urlopen(req) as r:
            raw = r.read()
            return json.loads(raw) if raw else {}

    def _refresh(self):
        try:
            runs = self._github_call(f"https://api.github.com/repos/{REPO}/actions/runs?per_page=1&status=in_progress")
            if runs.get("workflow_runs"): return self._json(429, {"ok": False, "error": "Already running"})
            self._github_call(f"https://api.github.com/repos/{REPO}/actions/workflows/deploy.yml/dispatches", method="POST", body=json.dumps({"ref":"main"}).encode())
            return self._json(200, {"ok": True, "message": "Refresh triggered"})
        except Exception as e:
            return self._json(500, {"ok": False, "error": str(e)[:200]})

    def _status(self):
        try:
            data = self._github_call(f"https://api.github.com/repos/{REPO}/actions/runs?per_page=1&event=push&status=completed")
            runs = data.get("workflow_runs", [])
            return self._json(200, {"conclusion": runs[0]["conclusion"] if runs else "unknown", "updated_at": runs[0]["updated_at"] if runs else None})
        except Exception as e:
            return self._json(500, {"error": str(e)[:200]})

    def _dqc_query(self):
        incoming = urllib.parse.parse_qs(urllib.parse.urlparse(self.path).query)
        q = {"limit": incoming.get("limit", ["1000"])[0]}
        for k in ("from", "to", "user", "status"):
            if incoming.get(k): q[k] = incoming[k][0]
        return urllib.parse.urlencode(q)

    def _fetch_dqc(self):
        if not DQC_API_KEY and not (DQC_USER and DQC_PASS):
            raise RuntimeError("DQC API key is not configured on the server")
        url = DQC_URL + "?" + self._dqc_query()
        headers = {"Accept": "application/json"}
        if DQC_API_KEY:
            headers["X-API-Key"] = DQC_API_KEY
        else:
            token = base64.b64encode(f"{DQC_USER}:{DQC_PASS}".encode()).decode()
            headers["Authorization"] = "Basic " + token
        req = urllib.request.Request(url, headers=headers)
        with urllib.request.urlopen(req, timeout=30) as r:
            data = json.loads(r.read().decode())
        raw_events = data.get("audits") or data.get("events") or []
        events = [normalize_dqc_event(e) for e in raw_events]
        events.sort(key=lambda e: e.get("ts", ""), reverse=True)
        meta = data.get("meta") if isinstance(data.get("meta"), dict) else {}
        return {
            "generated_at": meta.get("generated_at") or data.get("generated_at"),
            "source_generated_at": meta.get("source_generated_at"),
            "meta": meta,
            "people": data.get("people", []),
            "friction": data.get("friction", []),
            "ideas": data.get("ideas", []),
            "leaderboard": data.get("leaderboard", {}),
            "events": events,
            "summary": self._summarize(events),
        }

    def _summarize(self, events):
        verdicts = Counter((e.get("verdict") or "UNKNOWN").upper() for e in events)
        users = Counter(event_user(e) for e in events)
        return {"total_audits": len(events), "verdicts": dict(verdicts), "users": dict(users)}

    def _dqc_events(self):
        try: return self._json(200, self._fetch_dqc())
        except Exception as e: return self._json(500, {"error": str(e)[:200]})

    def _dqc_csv(self):
        try:
            data = self._fetch_dqc(); out = io.StringIO(); w = csv.writer(out)
            w.writerow(["date", "user", "order", "verdict", "rejection_reason", "timestamp_utc", "dqc_skill_version"])
            for e in data.get("events", []): w.writerow([(e.get("ts") or "")[:10], event_user(e), e.get("order",""), e.get("verdict",""), event_reason(e), e.get("ts",""), event_version(e)])
            return self._send(200, out.getvalue().encode(), "text/csv", "dqc_usage.csv")
        except Exception as e: return self._json(500, {"error": str(e)[:200]})

    def _dqc_xlsx(self):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            data = self._fetch_dqc(); events = data.get("events", []); summary = data.get("summary", {})
            wb = Workbook(); ws = wb.active; ws.title = "Runs"
            headers = ["date", "user", "order", "verdict", "rejection_reason", "timestamp_utc", "dqc_skill_version"]
            ws.append(headers)
            for e in events: ws.append([(e.get("ts") or "")[:10], event_user(e), e.get("order",""), e.get("verdict",""), event_reason(e), e.get("ts",""), event_version(e)])
            ws2 = wb.create_sheet("Summary"); ws2.append(["metric", "value"]); ws2.append(["total_audits", summary.get("total_audits", 0)]); ws2.append(["PASSED", summary.get("verdicts",{}).get("PASSED",0)]); ws2.append(["REJECTED", summary.get("verdicts",{}).get("REJECTED",0)]); ws2.append([]); ws2.append(["user", "audit_count"])
            for u,c in sorted(summary.get("users",{}).items(), key=lambda x: -x[1]): ws2.append([u,c])
            for sheet in (ws, ws2):
                for c in sheet[1]: c.font = Font(bold=True); c.fill = PatternFill("solid", fgColor="D9EAF7")
                for col in sheet.columns: sheet.column_dimensions[col[0].column_letter].width = min(max(len(str(cell.value or "")) for cell in col)+2, 45)
            bio = io.BytesIO(); wb.save(bio)
            return self._send(200, bio.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "dqc_usage.xlsx")
        except Exception as e: return self._json(500, {"error": str(e)[:200]})

    def _html(self, html, code=200): return self._send(code, html.encode(), "text/html; charset=utf-8")
    def _json(self, code, data): return self._send(code, json.dumps(data).encode(), "application/json")
    def _send(self, code, body, content_type, filename=None):
        self.send_response(code); self.send_header("Content-Type", content_type); self.send_header("Access-Control-Allow-Origin", "*"); self.send_header("Content-Length", str(len(body)))
        if filename: self.send_header("Content-Disposition", f'attachment; filename="{filename}"')
        self.end_headers(); self.wfile.write(body)
    def log_message(self, f, *a): pass

print(f"Server on :{PORT}", flush=True)
socketserver.TCPServer.allow_reuse_address = True
socketserver.TCPServer(("0.0.0.0", PORT), H).serve_forever()
